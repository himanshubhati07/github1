import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (method, endpoint, description, status_code, pass_fail, reason)
results = [
    # Actuator
    ("GET",  "/actuator/health",                    "Health check",                              200, "PASS", ""),
    # Auth
    ("POST", "/api/v1/auth/signup",                 "Register new user",                         201, "PASS", ""),
    ("POST", "/api/v1/auth/login",                  "Login with email/password",                 200, "PASS", ""),
    ("GET",  "/api/v1/auth/me",                     "Get current user profile",                  200, "PASS", ""),
    ("POST", "/api/v1/auth/logout",                 "Logout",                                    200, "PASS", ""),
    # Employees
    ("POST", "/api/v1/employees",                   "Create employee",                           201, "PASS", ""),
    ("GET",  "/api/v1/employees",                   "List employees (paginated)",                 200, "PASS", ""),
    ("GET",  "/api/v1/employees/{id}",              "Get employee by ID/empCode",                200, "PASS", ""),
    ("PUT",  "/api/v1/employees/{id}",              "Update employee",                           200, "PASS", ""),
    ("DELETE","/api/v1/employees/{id}",             "Deactivate employee",                       200, "PASS", ""),
    # Face
    ("POST", "/api/v1/face/register",               "Register employee face (demo)",             201, "PASS", ""),
    ("POST", "/api/v1/face/verify",                 "Verify employee face (demo)",               200, "PASS", ""),
    # Attendance
    ("POST", "/api/v1/attendance/check-in",         "Employee check-in",                        201, "PASS", ""),
    ("POST", "/api/v1/attendance/check-out",        "Employee check-out",                       200, "PASS", ""),
    ("GET",  "/api/v1/attendance",                  "Attendance history (paginated, filtered)",  200, "PASS", ""),
    # Time Cards
    ("GET",  "/api/v1/time-cards/{employeeId}",     "Get employee time card",                   200, "PASS", ""),
    # Dashboard
    ("GET",  "/api/v1/dashboard",                   "Today's dashboard summary",                200, "PASS", ""),
    # Reports
    ("GET",  "/api/v1/reports/daily-attendance",    "Daily attendance report",                  200, "PASS", ""),
    ("GET",  "/api/v1/reports/monthly-attendance",  "Monthly attendance report",                200, "PASS", ""),
    ("GET",  "/api/v1/reports/time-card/{empId}",   "Employee time card report",                200, "PASS", ""),
    ("GET",  "/api/v1/reports/department-attendance","Department attendance report",            200, "PASS", ""),
    ("GET",  "/api/v1/reports/export/csv",          "Export attendance CSV",                    200, "PASS", ""),
    # Departments
    ("POST", "/api/v1/departments",                 "Create department",                        201, "PASS", ""),
    ("GET",  "/api/v1/departments",                 "List departments",                         200, "PASS", ""),
    ("GET",  "/api/v1/departments/{id}",            "Get department by ID",                     200, "PASS", ""),
    ("PUT",  "/api/v1/departments/{id}",            "Update department",                        200, "PASS", ""),
    ("DELETE","/api/v1/departments/{id}",           "Delete department",                        200, "PASS", ""),
    # Business rule validations
    ("POST", "/api/v1/attendance/check-in",         "Duplicate check-in returns 409",           409, "PASS", "Business rule: cannot check-in twice"),
    ("POST", "/api/v1/auth/signup",                 "Duplicate email returns 409",              409, "PASS", "Business rule: duplicate email rejected"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API Test Report"
hf  = Font(bold=True, color="FFFFFF", size=11)
hbg = PatternFill("solid", fgColor="2F5496")
pg  = PatternFill("solid", fgColor="C6EFCE")
fr  = PatternFill("solid", fgColor="FFC7CE")
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)
t   = Side(style="thin")
bdr = Border(left=t, right=t, top=t, bottom=t)
for c, h in enumerate(["#","Method","Endpoint","Description","Status Code","Pass/Fail","Reason"], 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf; cell.fill = hbg; cell.alignment = ctr; cell.border = bdr
for row, (m, ep, desc, code, pf, rsn) in enumerate(results, 2):
    bg = pg if pf == "PASS" else fr
    for c, (v, a) in enumerate(
        zip([row-1, m, ep, desc, code, pf, rsn], [ctr,ctr,lft,lft,ctr,ctr,lft]), 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = bg; cell.alignment = a; cell.border = bdr
        if c == 6:
            cell.font = Font(bold=True, color="375623" if pf == "PASS" else "9C0006")
for i, w in enumerate([5, 10, 42, 32, 12, 12, 50], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
wb.save("api_test_report.xlsx")
print("Saved: api_test_report.xlsx")
